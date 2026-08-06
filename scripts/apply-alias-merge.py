#!/usr/bin/env python3
"""Apply docs/content/technique-alias-merge.csv to the technique spreadsheet.

    python3 scripts/apply-alias-merge.py <techniques.xlsx>

The other half of the procedure in docs/content/technique-alias-merge.md.
The sheet is the authoring surface for the 450 seeded techniques — a JSON edit
to one is reverted by the next re-import — so the aliases have to land here to
be durable.

ALL-OR-NOTHING, and that is the point. Every target row is located and checked
BEFORE anything is written: the row must exist, and its current `aliases` cell
must still equal the `current_aliases` column the CSV was generated against. A
single mismatch aborts with the row named and the file untouched.

Without that check this is a blind overwrite. The CSV holds a whole cell, not a
delta, so applying it to a sheet that has since gained an alias would silently
delete that alias — and nothing downstream would report a fault, because a
missing alias looks exactly like an alias nobody wrote.

Re-running after a successful apply is a no-op that FAILS the check rather than
succeeding quietly: the cells now hold `aliases_new_cell`, not
`current_aliases`. That is the correct behaviour — "already applied" and "drifted
since generation" are the same observation, and both want a human.
"""
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "docs/content/technique-alias-merge.csv"
SHEET = "Techniques"
FIRST_COL = "technique_id"


def slug(v) -> str:
    """The importer's id normalisation, so a sheet id in any casing matches."""
    return re.sub(r"[^a-z0-9]+", "-", str(v or "").strip().lower()).strip("-")


def norm(cell) -> str:
    """A cell's alias list as a comparable string.

    The sheet mixes `;` and `,` as separators within this column (see _split in
    import-exercise-catalog.py), so comparing raw text would report a spurious
    mismatch on a row that merely uses the other one.
    """
    return "; ".join(p.strip() for p in re.split(r"[;,]", str(cell or "")) if p.strip())


def main() -> None:
    if len(sys.argv) != 2:
        sys.exit(__doc__.strip().splitlines()[2].strip())
    path = Path(sys.argv[1])
    if not path.exists():
        sys.exit(f"no such spreadsheet: {path}")

    rows = list(csv.DictReader(CSV_PATH.open()))
    if not rows:
        sys.exit(f"{CSV_PATH.relative_to(ROOT)} is empty")

    # Not read_only: this workbook gets written back.
    wb = openpyxl.load_workbook(path)
    if SHEET not in wb.sheetnames:
        sys.exit(f"{path.name} has no {SHEET!r} sheet (found {wb.sheetnames})")
    ws = wb[SHEET]

    # The sheets carry a title banner above the real header row.
    header_row = None
    for r in ws.iter_rows(min_row=1, max_row=20):
        if r and str(r[0].value).strip() == FIRST_COL:
            header_row = r[0].row
            break
    if header_row is None:
        sys.exit(f"{path.name}: no {FIRST_COL!r} header in the first 20 rows")

    header = {}
    for c in ws[header_row]:
        if c.value:
            header[str(c.value).strip()] = c.column
    for col in (FIRST_COL, "aliases"):
        if col not in header:
            sys.exit(f"{path.name}: no {col!r} column")

    at_id = {}
    for row in range(header_row + 1, ws.max_row + 1):
        key = slug(ws.cell(row, header[FIRST_COL]).value)
        if not key:
            continue
        if key in at_id:
            # Silently taking the first would write one row and leave its twin
            # holding the old cell — a half-applied merge nobody can see.
            sys.exit(f"{path.name}: duplicate {FIRST_COL} {key!r} "
                     f"on rows {at_id[key]} and {row}")
        at_id[key] = row

    # Check everything first; write nothing until all 130 pass.
    planned, problems = [], []
    for r in rows:
        tid = r["technique_id"]
        row = at_id.get(tid)
        if row is None:
            problems.append(f"{tid}: not in the sheet")
            continue
        cell = ws.cell(row, header["aliases"])
        have = norm(cell.value)
        if have == norm(r["aliases_new_cell"]):
            problems.append(f"{tid}: already applied")
        elif have != norm(r["current_aliases"]):
            problems.append(
                f"{tid}: sheet has {have!r}, CSV was generated against "
                f"{norm(r['current_aliases'])!r}")
        else:
            planned.append((cell, r["aliases_new_cell"], r["aliases_to_add"]))

    if problems:
        print(f"REFUSED — {len(problems)} of {len(rows)} rows did not check out; "
              f"{path.name} is unchanged.\n", file=sys.stderr)
        for p in problems[:25]:
            print(f"  {p}", file=sys.stderr)
        if len(problems) > 25:
            print(f"  ... and {len(problems) - 25} more", file=sys.stderr)
        sys.exit(1)

    for cell, new_value, _ in planned:
        cell.value = new_value
    wb.save(path)

    added = sum(len(
        [x for x in a.split(";") if x.strip()]) for _, _, a in planned)
    print(f"{path.name}: {len(planned)} rows updated, {added} aliases added")


if __name__ == "__main__":
    main()
